from io import BytesIO
import pandas as pd
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill, Font

# Define colors
LIGHT_BLUE = PatternFill(start_color="DAE9F8", end_color="DAE9F8", fill_type="solid")
LIGHT_ORANGE = PatternFill(start_color="FBE2D5", end_color="FBE2D5", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")
ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# Format
NUMBER_FORMAT = "#,##0.00"

# Define the numeric columns for each Shopify sheet
NUMERIC_COLS = {
    "Shopify payments": [
        "Transaction ID", "Gross payments", "Refunds", "Net payments"
    ],
    "Shopify incl. returns": [
        "Order ID", "Sale ID",
        "Gross sales", "Discounts", "Returns", "Net sales",
        "Shipping", "Taxes", "Total sales", "Net quantity"
    ],
    "Shopify Tax": [
        "Sale tax ID", "Order ID", "Amount", "Rate"
    ],
    "ITSP Returns": [
        "Return costs", "Discount", "Amount", "Postage costs", "Shipping cost original order", "Shipping cost return", "VAT %", "Total EUR incl. VAT", "Check"
    ],
    "Old ITSP": [
        "Shipping costs", "Discount", "Amount", "VAT value", "Payment amount (LCY)", "Total Qty", "Subtotaal excl VAT", "Total incl. VAT", "VAT %"
    ]
}

DATE_COLS = {
    "Shopify payments": ["Date"],
    "Shopify incl. returns": ["Date"],
    "Shopify Tax": ["Date"],
    "ITSP Sales": ["Date"],
    "ITSP Returns": ["Date"],
}

def export_to_excel(sheets: dict):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            if df.empty:
                df = pd.DataFrame(["No data"], columns=["Info"])

            # -------------------
            # Clean numeric columns
            # -------------------
            for col in NUMERIC_COLS.get(sheet, []):
                if col in df.columns:
                    # Convert EU-style numbers to float
                    df[col] = df[col].astype(str).str.replace(",", ".")
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype(float)

            # -------------------
            # Clean date columns
            # -------------------
            for col in DATE_COLS.get(sheet, []):
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors="coerce")

            # Write to Excel
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.book[sheet]

            # -------------------
            # Apply sheet-specific columns
            # -------------------
            if sheet == "Shopify incl. returns":
                add_shopify_returns_columns(ws)
            elif sheet == "Shopify payments":
                add_shopify_payments_columns(ws)
            elif sheet == "ITSP Sales":
                add_itsp_sales_columns(ws)
            elif sheet == "ITSP Returns":
                add_itsp_returns_columns(ws)

            # -------------------
            # Apply formatting
            # -------------------
            # for col_idx, col_name in enumerate(df.columns, start=1):
            #     if col_name in NUMERIC_COLS.get(sheet, []):
            #         for row in range(2, ws.max_row + 1):
            #             ws.cell(row=row, column=col_idx).number_format = NUMBER_FORMAT
            #     if col_name in DATE_COLS.get(sheet, []):
            #         for row in range(2, ws.max_row + 1):
            #             ws.cell(row=row, column=col_idx).number_format = "dd/mm/yyyy"
            del df
            import gc
            gc.collect()
            ws.auto_filter.ref = ws.dimensions

        # -------------------
        # Add reconciliation sheet
        # -------------------
        add_reconciliation_sheet(writer.book)

        # -------------------
        # Reorder and color tabs
        # -------------------
        desired_order = ["Recon", "ITSP Sales", "ITSP Returns", "Shopify incl. returns",
                         "Old ITSP", "Shopify payments", "Shopify Tax", "Backend"]
        writer.book._sheets.sort(key=lambda ws: desired_order.index(ws.title) if ws.title in desired_order else 999)
        color_sheet_tabs(writer.book)

    output.seek(0)
    return output

def color_sheet_tabs(wb):
    TAB_COLORS = {
        "ITSP Sales": "DAF2D0",
        "ITSP Returns": "DAF2D0",
        "Shopify incl. returns": "DAF2D0",
        "Old ITSP": "FBE2D5",
        "Shopify payments": "FBE2D5",
        "Shopify Tax": "DAE9F8",
    }

    for sheet_name, color in TAB_COLORS.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

def add_reconciliation_sheet(wb):
    ws = wb.create_sheet("Recon")

    headers = [
        "Order Ref", "Date", "Country", "VAT %",
        "VAT % (Old)", "Diff", "In ITSP?",
        "Cancelled", "Gift card", "Gift card 2",
        "ITSP Sales", "ITSP Return", "Total ITSP",
        "Shopify Sales", "Shopify Return",
        "Total Shopify", "Delta", "Comment"
    ]

    # Headers in row 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
        ws.cell(row=4, column=col).font = Font(bold=True)
        if h == "Order Ref":
            ws.cell(row=4, column=col).fill = LIGHT_GREEN

    def extract_orders(sheet_name, column_name):
        """Extract non-empty Order values from a sheet"""
        sheet = wb[sheet_name]

        order_col = None
        for c in range(1, sheet.max_column + 1):
            if sheet.cell(1, c).value == column_name:
                order_col = c
                break

        if order_col is None:
            raise ValueError(f"Order column not found in {sheet_name}")

        return {
            sheet.cell(r, order_col).value
            for r in range(2, sheet.max_row + 1)
            if sheet.cell(r, order_col).value
        }

    # ✅ Union of orders from both sources
    shopify_orders = extract_orders("Shopify incl. returns", "Order")
    itsp_orders = extract_orders("ITSP Sales", "Reference")
    itsp_returns = extract_orders("ITSP Returns","Comments")

    all_orders = sorted(shopify_orders | itsp_orders | itsp_returns)

    # Write orders starting at row 5
    for row_idx, order in enumerate(all_orders, start=5):
        ws.cell(row=row_idx, column=1, value=order)
        # ws.cell(row_idx, column=1).fill = LIGHT_GREEN

    last_row = 5 + len(all_orders) - 1
    fill_reconciliation_formulas(ws, last_row)

def fill_reconciliation_formulas(ws, last_row):
    # Sum headers
    sum_cols = ["K", "L", "M", "N", "O", "P"]
    for col in sum_cols:
        cell_str = f"{col}1"
        ws[cell_str] = f"=SUM({col}5:{col}{last_row})"

        if col[0] in ["K", "L", "M"]:
            ws[cell_str].fill = LIGHT_BLUE
            ws[cell_str].font = Font(bold=True)
        if col[0] in ["N", "O", "P"]:
            ws[cell_str].fill = LIGHT_ORANGE
            ws[cell_str].font = Font(bold=True)
    
    ws["Q1"] = f"=SUBTOTAL(9,Q5:Q{last_row})"
    ws["Q1"].fill = ORANGE
    ws["Q1"].font = Font(bold=True)

    # Static values
    ws["J2"] = 2024
    ws["J2"].font = Font(bold=True)
    ws["J3"] = 4
    ws["J3"].font = Font(bold=True)
    ws["N2"] = "order"
    ws["N2"].font = Font(bold=True)
    ws["N3"] = "Shopify incl. VAT"
    ws["N3"].font = Font(bold=True)
    ws["O2"] = "return"
    ws["O2"].font = Font(bold=True)

    formulas = {
        "B5": '=IFERROR(IFERROR(VLOOKUP(A5,\'Shopify incl. returns\'!C:D,2,0),VLOOKUP(A5,\'ITSP Returns\'!H:P,9,0)),VLOOKUP(A5,\'ITSP Sales\'!F:AB,23,0))',
        "C5": '=IFERROR(VLOOKUP(A5,\'Shopify incl. returns\'!C:X,22,0),VLOOKUP(A5,\'Old ITSP\'!F:G,2,0))',
        "D5": '=IFERROR(AVERAGEIFS(\'Shopify Tax\'!N:N,\'Shopify Tax\'!F:F,A5),0)',
        "E5": '=IFERROR(VLOOKUP(A5,\'Old ITSP\'!F:Z,21,0),D5)',
        "F5": '=D5-E5',
        "G5": '=IF(A5=IFERROR(VLOOKUP(A5,\'Old ITSP\'!F:F,1,0),0),"Yes","No")',
        "H5": '=IFERROR(IF(VLOOKUP(A5,\'Old ITSP\'!F:L,7,0)="Canceled","Canceled",""),"")',
        "I5": '=SUMIFS(\'Shopify payments\'!I:I,\'Shopify payments\'!B:B,A5,\'Shopify payments\'!C:C,$I$4)',
        "J5": '=SUMIFS(\'Shopify payments\'!I:I,\'Shopify payments\'!J:J,$J$2,\'Shopify payments\'!K:K,$J$3,\'Shopify payments\'!B:B,A5,\'Shopify payments\'!C:C,$I$4)',
        "K5": '=SUMIFS(\'ITSP Sales\'!Y:Y,\'ITSP Sales\'!F:F,A5)',
        "L5": '=SUMIFS(\'ITSP Returns\'!R:R,\'ITSP Returns\'!H:H,A5)*-1',
        "M5": '=ROUND(SUM(K5:L5),2)',
        "N5": '=SUMIFS(\'Shopify incl. returns\'!$V:$V,\'Shopify incl. returns\'!$C:$C,$A5,\'Shopify incl. returns\'!$E:$E,N$2)',
        "O5": '=SUMIFS(\'Shopify incl. returns\'!$V:$V,\'Shopify incl. returns\'!$C:$C,$A5,\'Shopify incl. returns\'!$E:$E,O$2)',
        "P5": '=ROUND(SUM(N5:O5),2)',
        "Q5": '=M5-P5',
    }

    # Fill formulas for all rows efficiently
    for r in range(5, last_row + 1):
        for col, f in formulas.items():
            cell_str = f"{col[0]}{r}"
            ws[cell_str].value = Translator(f, origin=col).translate_formula(dest=f"{col[0]}{r}")

            # If this is a date column (e.g., B), apply date format
            if col[0] in ["B"]:
                ws[cell_str].number_format = "dd/mm/yyyy"

            # if col[0] in ["K", "L", "M"]:
            #     ws[cell_str].fill = LIGHT_BLUE
            
            # if col[0] in ["N", "O", "P"]:
            #     ws[cell_str].fill = LIGHT_ORANGE
            
            # if col[0] in ["Q"]:
            #     ws[cell_str].fill = ORANGE

def add_shopify_returns_columns(ws):
    last_col = ws.max_column
    last_row = ws.max_row

    check_col = last_col + 1
    country_code_col = last_col + 2

    ws.cell(1, check_col, "CHECK")
    ws.cell(1, check_col).fill = LIGHT_ORANGE
    ws.cell(1, country_code_col, "Country code")
    ws.cell(1, country_code_col).fill = LIGHT_BLUE

    for r in range(2, last_row + 1):
        ws.cell(r, check_col).fill = LIGHT_ORANGE
        ws.cell(r, country_code_col).fill = LIGHT_BLUE

        ws.cell(r, check_col).value = f"=SUM(S{r}:U{r})-V{r}"
        ws.cell(r, country_code_col).value = (
            f'=IF(I{r}="",'
            f'VLOOKUP(H{r},Backend!E:F,2,0),'
            f'VLOOKUP(I{r},Backend!E:F,2,0))'
        )
    
    # header_font = Font(bold=True)

    # for col in range(1, ws.max_column + 1):
    #     ws.cell(row=1, column=col).font = header_font

def add_shopify_payments_columns(ws):
    last_col = ws.max_column
    last_row = ws.max_row

    year_col = last_col + 1
    month_col = last_col + 2

    ws.cell(1, year_col, "Year")
    ws.cell(1, year_col).fill = LIGHT_GREEN
    ws.cell(1, month_col, "Month")
    ws.cell(1, month_col).fill = LIGHT_GREEN

    for r in range(2, last_row + 1):
        ws.cell(r, year_col).value = f"=YEAR(B{r})"
        ws.cell(r, month_col).value = f"=MONTH(B{r})"
    
    # header_font = Font(bold=True)

    # for col in range(1, ws.max_column + 1):
    #     ws.cell(row=1, column=col).font = header_font

def add_itsp_sales_columns(ws):
    last_col = ws.max_column
    last_row = ws.max_row

    total_col = last_col + 1
    vat_col = last_col + 2
    date_col = last_col + 3

    ws.cell(1, total_col, "Total EUR incl. VAT")
    ws.cell(1, total_col).fill = LIGHT_BLUE
    ws.cell(1, vat_col, "VAT %")
    ws.cell(1, vat_col).fill = LIGHT_BLUE
    ws.cell(1, date_col, "Date")
    ws.cell(1, date_col).fill = LIGHT_BLUE

    for r in range(2, last_row + 1):
        # # Colors
        # ws.cell(r, total_col).fill = LIGHT_BLUE
        # ws.cell(r, vat_col).fill = LIGHT_BLUE
        # ws.cell(r, date_col).fill = LIGHT_BLUE
        # Formulas
        ws.cell(r, total_col).value = f"=H{r}+P{r}+Q{r}"
        ws.cell(r, vat_col).value = f"=Q{r}/(H{r}+P{r})"
        ws.cell(r, date_col).value = f"=B{r}"
        # Format
        ws.cell(r, total_col).number_format = NUMBER_FORMAT
        ws.cell(r, vat_col).number_format = NUMBER_FORMAT
        ws.cell(r, date_col).number_format = "dd/mm/yyyy"
    
    # header_font = Font(bold=True)

    # for col in range(1, ws.max_column + 1):
    #     ws.cell(row=1, column=col).font = header_font


# --------------------------------------------------
# ITSP Returns extra columns
# --------------------------------------------------
def add_itsp_returns_columns(ws):
    last_col = ws.max_column
    last_row = ws.max_row

    date_col = last_col + 1
    ship_cost_col = last_col + 2
    ship_cost_return_col = last_col + 3
    vat_col = last_col + 4
    total_col = last_col + 5
    check_col = last_col + 6

    ws.cell(1, date_col, "Date")
    ws.cell(1, date_col).fill = LIGHT_BLUE
    ws.cell(1, ship_cost_col, "Shipping cost original order")
    ws.cell(1, ship_cost_col).fill = LIGHT_BLUE
    ws.cell(1, ship_cost_return_col, "Shipping cost return")
    ws.cell(1, ship_cost_return_col).fill = LIGHT_BLUE
    ws.cell(1, vat_col, "VAT %")
    ws.cell(1, vat_col).fill = LIGHT_BLUE
    ws.cell(1, total_col, "Total EUR incl. VAT")
    ws.cell(1, total_col).fill = LIGHT_BLUE
    ws.cell(1, check_col, "Check")
    ws.cell(1, check_col).fill = LIGHT_ORANGE

    for r in range(2, last_row + 1):
        # Colors
        # ws.cell(r, date_col).fill = LIGHT_BLUE
        # ws.cell(r, ship_cost_col).fill = LIGHT_BLUE
        # ws.cell(r, ship_cost_return_col).fill = LIGHT_BLUE
        # ws.cell(r, vat_col).fill = LIGHT_BLUE
        # ws.cell(r, total_col).fill = LIGHT_BLUE
        # ws.cell(r, check_col).fill = LIGHT_ORANGE

        # Formulas
        ws.cell(r, date_col).value = f"=B{r}"
        ws.cell(r, ship_cost_col).value = f"=VLOOKUP(H{r},'Old ITSP'!F:H,3,0)"
        ws.cell(r, ship_cost_return_col).value = f"=IF(K{r}=SUMIFS('Old ITSP'!W:W,'Old ITSP'!F:F,H{r}),O{r},0)"
        ws.cell(r, vat_col).value = f"=ROUND(VLOOKUP(H{r},'Old ITSP'!F:Z,21,0),2)"
        ws.cell(r, total_col).value = f"=(L{r}+P{r})*(1+Q{r})"
        ws.cell(r, check_col).value = f"=VLOOKUP(H{r},'ITSP Sales'!F:F,1,0)"

        # Format
        ws.cell(r, date_col).number_format = "dd/mm/yyyy"
        # ws.cell(r, ship_cost_col).number_format = NUMBER_FORMAT
        # ws.cell(r, ship_cost_return_col).number_format = NUMBER_FORMAT
        # ws.cell(r, vat_col).number_format = NUMBER_FORMAT
        # ws.cell(r, total_col).number_format = NUMBER_FORMAT
    
    # header_font = Font(bold=True)

    # for col in range(1, ws.max_column + 1):

    #     ws.cell(row=1, column=col).font = header_font



